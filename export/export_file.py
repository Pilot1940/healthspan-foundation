"""export/export_file.py — xlsx and PDF export of any lib/views view.

Consumes lib/views ONLY. openpyxl / reportlab are imported lazily so the module
loads even if they're absent (the skill degrades to markdown).

  to_xlsx(conn, names, profile_id, out_path, **params)  -> path  (one sheet per view + a summary sheet)
  to_pdf(conn, names, profile_id, out_path, **params)   -> path  (formatted report)
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from lib.views import run_view


def _stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _cell(v):
    """Coerce a value to something Excel/PDF can render (lists/dicts → string)."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v)
    return str(v)


def to_xlsx(conn, names, profile_id: str, out_path: str, **params) -> str:
    """Write one worksheet per view (raw rows) + a Summary sheet listing each view's summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append([f"HealthSpan export — {_stamp()}"])
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.append([])

    for name in names:
        res = run_view(conn, name, profile_id, **params)
        # raw sheet (Excel sheet names max 31 chars, no special chars)
        sheet = name[:31].replace("/", "_")
        ws = wb.create_sheet(sheet)
        cols = res["columns"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        for row in res["rows"]:
            ws.append([_cell(row.get(c)) for c in cols])
        # summary line on the Summary sheet
        ws_sum.append([name, res["description"]])
        if res.get("summary"):
            for k, v in res["summary"].items():
                ws_sum.append(["", k, _cell(v)])

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def to_pdf(conn, names, profile_id: str, out_path: str, **params) -> str:
    """Formatted multi-section PDF: title + one table per view (+ summary rows)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(out_path, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    el = [Paragraph("HealthSpan Report", styles["Title"]),
          Paragraph(_stamp(), styles["Normal"]), Spacer(1, 0.5 * cm)]

    for name in names:
        res = run_view(conn, name, profile_id, **params)
        el.append(Paragraph(name, styles["Heading2"]))
        el.append(Paragraph(res["description"], styles["Italic"]))
        if res.get("summary"):
            summ = " · ".join(f"{k}: {v}" for k, v in res["summary"].items())
            el.append(Paragraph(summ, styles["Normal"]))
        el.append(Spacer(1, 0.2 * cm))
        cols = res["columns"]
        rows = res["rows"][:60]  # cap for PDF
        if rows:
            data = [cols] + [[("" if r.get(c) is None else str(_cell(r.get(c)))) for c in cols] for r in rows]
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f7")]),
            ]))
            el.append(t)
        else:
            el.append(Paragraph("<i>no rows</i>", styles["Normal"]))
        el.append(Spacer(1, 0.5 * cm))

    doc.build(el)
    return out_path


if __name__ == "__main__":
    import argparse, os
    sys.path.insert(0, "scripts"); import hs_ops; hs_ops._load_env()
    import psycopg2
    ap = argparse.ArgumentParser()
    ap.add_argument("--format", choices=["xlsx", "pdf"], required=True)
    ap.add_argument("--views", default="v_recovery_30d,abnormal_labs,sprints_status")
    ap.add_argument("--profile", default="21f69003-46f8-4e1c-a928-b1f694ce4aff")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    names = [n.strip() for n in args.views.split(",")]
    fn = to_xlsx if args.format == "xlsx" else to_pdf
    print("Wrote:", fn(conn, names, args.profile, args.out))
    conn.close()
